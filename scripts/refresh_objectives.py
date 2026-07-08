#!/usr/bin/env python3
"""
Refresh references/learning-objectives.json from the 720 management Excel file.

The management file has two ID tabs — "ID מתמטיקה" and "ID מדעים" — each
listing learning objectives in order. This script reads the unit-level IDs
(rows with a full ID like `methodica-<subject>-<topic>-NN`) and saves them
to a JSON file next to the skill's references.

Run this whenever new objectives are added to the management file.

Usage:
    python refresh_objectives.py <path-to-management-xlsx>
"""

import sys
import io
import re
import json
import argparse
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except Exception:
        pass

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit('openpyxl not installed. Run: pip install openpyxl')


# Unit-level pattern: methodica-<letters/dashes>-<NN>, no more segments after NN
UNIT_ID_RE = re.compile(r'^methodica-[a-zA-Z-]+-\d{1,2}$')


def extract_tab(ws):
    """Extract ordered unit IDs from a tab."""
    entries = []
    seen = set()
    for row in ws.iter_rows(values_only=True):
        # Row layout in ID tabs: נושא | מס' | יעד | ID | ...
        topic, num, obj_text, unit_id = row[0], row[1], row[2], row[3]
        if not (unit_id and isinstance(unit_id, str) and UNIT_ID_RE.match(unit_id)):
            continue
        if unit_id in seen:
            continue
        seen.add(unit_id)
        entries.append({
            'id': unit_id,
            'number': num,
            'topic': topic,
            'objective': obj_text[:200] if obj_text else '',
        })
    return entries


def main():
    ap = argparse.ArgumentParser(description='Refresh learning objectives from 720 management Excel.')
    ap.add_argument('xlsx', help='Path to management Excel file (קובץ ניהול 720 תשפז.xlsx)')
    ap.add_argument('--out', default=None, help='Output JSON path (default: ../references/learning-objectives.json)')
    args = ap.parse_args()

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = Path(__file__).parent.parent / 'references' / 'learning-objectives.json'

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)

    result = {}
    for tab_name, key in [('ID מתמטיקה', 'math'), ('ID מדעים', 'science')]:
        if tab_name not in wb.sheetnames:
            print(f'WARNING: tab "{tab_name}" not found in workbook', file=sys.stderr)
            result[key] = []
            continue
        result[key] = extract_tab(wb[tab_name])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open('w', encoding='utf-8', newline='\n') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    for k, entries in result.items():
        print(f'{k}: {len(entries)} objectives')
    print(f'Saved: {out_path}')


if __name__ == '__main__':
    main()
