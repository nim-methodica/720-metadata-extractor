#!/usr/bin/env python3
"""
Refresh the MOE learning-objective index from a ministry-provided Excel file.

The ministry provides one Excel per subject+grade (e.g., "אינדקס יעדי למידה - מדעים
(כימיה) כיתה ז׳"). This script:

1. Reads the "יעדי למידה" sheet from ONE Excel file.
2. Detects subject (math / science) from the code prefix in column E.
3. Merges the entries into references/moe-index.json (replacing that subject's list).
4. Re-runs auto-matching against references/learning-objectives.json,
   updating moe_code and subtopic_code for matched entries.

Usage:
    python refresh_moe_index.py "<path/to/index.xlsx>"

If you have multiple index files, run this script once per file.
"""

import sys
import io
import re
import json
import argparse
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except Exception:
        pass

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit('openpyxl not installed. Run: pip install openpyxl')


def normalize(s):
    if not s: return ''
    s = str(s)
    s = re.sub(r'[.,;:\-–—"\'()\[\]{}־׳״]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def similarity(a, b):
    a_tokens = set(normalize(a).split())
    b_tokens = set(normalize(b).split())
    if not a_tokens or not b_tokens:
        return 0
    return len(a_tokens & b_tokens) / max(len(a_tokens), len(b_tokens))


def extract_index(xlsx_path: Path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if 'יעדי למידה' not in wb.sheetnames:
        raise SystemExit(f'Sheet "יעדי למידה" not found. Available: {wb.sheetnames}')
    ws = wb['יעדי למידה']

    entries = []
    cd, ct, cst = None, None, None
    for row in ws.iter_rows(values_only=True, min_row=2):
        vals = list(row) + [None] * (5 - len(row))
        domain, topic, subtopic, objective, code = vals[:5]
        if domain: cd = domain
        if topic: ct = topic
        if subtopic: cst = subtopic
        if not (objective and code):
            continue
        entries.append({
            'domain_he': cd,
            'topic_he': ct,
            'subtopic_he': cst,
            'objective_he': str(objective).strip(),
            'code': str(code).strip(),
        })
    return entries


def detect_subject(entries):
    """Detect subject (math or science) from the code prefix."""
    for e in entries:
        code = e.get('code', '')
        if code.startswith('MOE.MATH'):
            return 'math'
        if code.startswith('MOE.SCI'):
            return 'science'
    return None


def rematch(objectives, moe_index, threshold=0.35):
    """Update moe_code / subtopic_code for methodica objectives based on text similarity."""
    for subject, entries in objectives.items():
        moe_entries = moe_index.get(subject, [])
        for mth in entries:
            # Reset any previous match
            mth.pop('moe_code', None)
            mth.pop('subtopic_code', None)
            mth.pop('subtopic_he', None)
            mth.pop('topic_he_moe', None)

            best, best_score = None, 0
            for m in moe_entries:
                s = similarity(mth.get('objective', ''), m.get('objective_he', ''))
                if s > best_score:
                    best_score, best = s, m
            if best and best_score >= threshold:
                mth['moe_code'] = best['code']
                mth['topic_he_moe'] = best['topic_he']
                mth['subtopic_he'] = best['subtopic_he']
                parts = best['code'].split('.')
                if len(parts) >= 6:
                    mth['subtopic_code'] = '.'.join(parts[:6])
    return objectives


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', help='Path to a ministry index Excel file')
    ap.add_argument('--index', default=None, help='Path to moe-index.json (default: skill references)')
    ap.add_argument('--objectives', default=None,
                    help='Path to learning-objectives.json (default: skill references)')
    args = ap.parse_args()

    ref = Path(__file__).parent.parent / 'references'
    idx_path = Path(args.index) if args.index else ref / 'moe-index.json'
    obj_path = Path(args.objectives) if args.objectives else ref / 'learning-objectives.json'

    # Extract from Excel
    entries = extract_index(Path(args.xlsx))
    subject = detect_subject(entries)
    if subject is None:
        raise SystemExit('Could not detect subject (math/science) from codes in the file.')
    print(f'Loaded {len(entries)} entries — subject: {subject}')

    # Merge into moe-index.json
    if idx_path.exists():
        with open(idx_path, encoding='utf-8') as f:
            moe_index = json.load(f)
    else:
        moe_index = {}
    moe_index[subject] = entries
    with open(idx_path, 'w', encoding='utf-8', newline='\n') as f:
        json.dump(moe_index, f, ensure_ascii=False, indent=2)
    print(f'Updated: {idx_path}')

    # Re-match
    with open(obj_path, encoding='utf-8') as f:
        objectives = json.load(f)
    objectives = rematch(objectives, moe_index)
    with open(obj_path, 'w', encoding='utf-8', newline='\n') as f:
        json.dump(objectives, f, ensure_ascii=False, indent=2)

    matched = sum(1 for e in objectives.get(subject, []) if 'moe_code' in e)
    total = len(objectives.get(subject, []))
    print(f'Rematched: {matched}/{total} {subject} objectives now have MOE codes')


if __name__ == '__main__':
    main()
